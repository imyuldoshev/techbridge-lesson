import datetime

from django import forms as django_forms
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from .forms import LessonForm, TeacherForm
from .models import Lesson, Teacher

WEEKDAY_ORDER = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']


def _is_admin(user):
    return user.is_superuser or user.is_staff


def _is_ustoz(user):
    return user.groups.filter(name='ustoz').exists()


def _is_bola(user):
    return not _is_admin(user) and not _is_ustoz(user)


def _get_teacher(user):
    return getattr(user, 'teacher_profile', None)


@login_required
def index(request):
    if _is_bola(request.user):
        return redirect('schedule')

    today_index = datetime.date.today().weekday()
    today_key = WEEKDAY_ORDER[today_index]
    tomorrow_key = WEEKDAY_ORDER[(today_index + 1) % 7]

    teacher = _get_teacher(request.user)

    if _is_admin(request.user):
        bugungi_darslar = Lesson.objects.filter(hafta_kuni=today_key, faol=True).select_related('ustoz')
        ertangi_darslar = Lesson.objects.filter(hafta_kuni=tomorrow_key, faol=True).select_related('ustoz')
        context = {
            'bugungi_darslar': bugungi_darslar,
            'ertangi_darslar': ertangi_darslar,
            'jami_ustozlar': Teacher.objects.count(),
            'jami_darslar': Lesson.objects.count(),
            'faol_darslar': Lesson.objects.filter(faol=True).count(),
        }
    else:
        own_qs = Lesson.objects.filter(ustoz=teacher) if teacher else Lesson.objects.none()
        bugungi_darslar = own_qs.filter(hafta_kuni=today_key, faol=True).select_related('ustoz')
        ertangi_darslar = own_qs.filter(hafta_kuni=tomorrow_key, faol=True).select_related('ustoz')
        context = {
            'bugungi_darslar': bugungi_darslar,
            'ertangi_darslar': ertangi_darslar,
            'jami_darslar': own_qs.count(),
            'faol_darslar': own_qs.filter(faol=True).count(),
        }

    return render(request, 'schedule/index.html', context)


@login_required
def teacher_list(request):
    if not _is_admin(request.user):
        return redirect('index')

    qs = Teacher.objects.annotate(darslar_soni=Count('darslar'))
    qidiruv = request.GET.get('q', '').strip()
    if qidiruv:
        qs = qs.filter(ism__icontains=qidiruv)
    mutaxassislik = request.GET.get('mutaxassislik', '').strip()
    if mutaxassislik:
        qs = qs.filter(mutaxassislik=mutaxassislik)
    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    mutaxassisliklar = Teacher.objects.exclude(mutaxassislik__isnull=True).exclude(
        mutaxassislik=''
    ).values_list('mutaxassislik', flat=True).distinct()
    context = {
        'page_obj': page_obj,
        'qidiruv': qidiruv,
        'mutaxassislik': mutaxassislik,
        'mutaxassisliklar': mutaxassisliklar,
    }
    return render(request, 'schedule/teacher_list.html', context)


@login_required
def teacher_form_view(request, pk=None):
    if not _is_admin(request.user):
        return redirect('index')

    teacher = get_object_or_404(Teacher, pk=pk) if pk else None
    if request.method == 'POST':
        form = TeacherForm(request.POST, instance=teacher)
        if form.is_valid():
            form.save()
            messages.success(request, "Ustoz muvaffaqiyatli saqlandi.")
            return redirect('teacher_list')
    else:
        form = TeacherForm(instance=teacher)
    return render(request, 'schedule/teacher_form.html', {'form': form, 'teacher': teacher})


@login_required
def teacher_delete(request, pk):
    if not _is_admin(request.user):
        return redirect('index')

    teacher = get_object_or_404(Teacher, pk=pk)
    if request.method == 'POST':
        teacher.delete()
        messages.success(request, "Ustoz o'chirildi.")
        return redirect('teacher_list')
    return render(request, 'schedule/teacher_confirm_delete.html', {'teacher': teacher})


@login_required
def lesson_list(request):
    if _is_bola(request.user):
        return redirect('schedule')

    teacher = _get_teacher(request.user)
    qs = Lesson.objects.select_related('ustoz')

    if _is_ustoz(request.user) and teacher:
        qs = qs.filter(ustoz=teacher)

    qidiruv = request.GET.get('q', '').strip()
    if qidiruv:
        qs = qs.filter(
            Q(fan_nomi__icontains=qidiruv) | Q(mavzu__icontains=qidiruv) | Q(ustoz__ism__icontains=qidiruv)
        )
    hafta_kuni = request.GET.get('hafta_kuni', '').strip()
    if hafta_kuni:
        qs = qs.filter(hafta_kuni=hafta_kuni)

    ustoz_id = ''
    if _is_admin(request.user):
        ustoz_id = request.GET.get('ustoz', '').strip()
        if ustoz_id:
            qs = qs.filter(ustoz_id=ustoz_id)

    faol = request.GET.get('faol', '').strip()
    if faol in ('1', '0'):
        qs = qs.filter(faol=(faol == '1'))

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj,
        'qidiruv': qidiruv,
        'hafta_kuni': hafta_kuni,
        'ustoz_id': ustoz_id,
        'faol': faol,
        'hafta_kuni_choices': Lesson.HAFTA_KUNI_CHOICES,
        'ustozlar': Teacher.objects.all() if _is_admin(request.user) else Teacher.objects.none(),
    }
    return render(request, 'schedule/lesson_list.html', context)


@login_required
def lesson_form_view(request, pk=None):
    if _is_bola(request.user):
        return redirect('schedule')

    teacher = _get_teacher(request.user)
    lesson = get_object_or_404(Lesson, pk=pk) if pk else None

    if lesson and _is_ustoz(request.user) and teacher and lesson.ustoz != teacher:
        messages.error(request, "Siz faqat o'z darslaringizni tahrirlashingiz mumkin.")
        return redirect('lesson_list')

    if request.method == 'POST':
        form = LessonForm(request.POST, instance=lesson)
        if _is_ustoz(request.user) and teacher:
            form.fields['ustoz'].widget = django_forms.HiddenInput()
            form.fields['ustoz'].initial = teacher
        if form.is_valid():
            saved = form.save(commit=False)
            if _is_ustoz(request.user) and teacher:
                saved.ustoz = teacher
            saved.save()
            messages.success(request, "Dars muvaffaqiyatli saqlandi.")
            return redirect('lesson_list')
    else:
        form = LessonForm(instance=lesson)
        if _is_ustoz(request.user) and teacher:
            form.fields['ustoz'].widget = django_forms.HiddenInput()
            form.fields['ustoz'].initial = teacher

    return render(request, 'schedule/lesson_form.html', {'form': form, 'lesson': lesson})


@login_required
def lesson_delete(request, pk):
    if _is_bola(request.user):
        return redirect('schedule')

    teacher = _get_teacher(request.user)
    lesson = get_object_or_404(Lesson, pk=pk)

    if _is_ustoz(request.user) and teacher and lesson.ustoz != teacher:
        messages.error(request, "Siz faqat o'z darslaringizni o'chirishingiz mumkin.")
        return redirect('lesson_list')

    if request.method == 'POST':
        lesson.delete()
        messages.success(request, "Dars o'chirildi.")
        return redirect('lesson_list')
    return render(request, 'schedule/lesson_confirm_delete.html', {'lesson': lesson})


@login_required
def lesson_export_excel(request):
    if _is_bola(request.user):
        return redirect('schedule')

    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Darslar"

    headers = ['T/r', 'Fan', 'Ustoz', 'Kun', 'Boshlanish', 'Tugash', 'Xona', 'Guruh', 'Holat']
    ws.append(headers)

    teacher = _get_teacher(request.user)
    qs = Lesson.objects.select_related('ustoz').all()
    if _is_ustoz(request.user) and teacher:
        qs = qs.filter(ustoz=teacher)

    for idx, lesson in enumerate(qs, start=1):
        ws.append([
            idx,
            lesson.fan_nomi,
            lesson.ustoz.ism,
            lesson.get_hafta_kuni_display(),
            lesson.boshlanish_vaqti.strftime('%H:%M'),
            lesson.tugash_vaqti.strftime('%H:%M'),
            lesson.xona or '',
            lesson.guruh or '',
            'Faol' if lesson.faol else 'Faol emas',
        ])

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="darslar.xlsx"'
    wb.save(response)
    return response


@login_required
def schedule_view(request):
    hafta_turi = request.GET.get('hafta_turi', 'toq').strip()
    qs = Lesson.objects.filter(faol=True).select_related('ustoz')

    teacher = _get_teacher(request.user)
    ustoz_id = request.GET.get('ustoz', '').strip()

    if _is_bola(request.user):
        if ustoz_id:
            qs = qs.filter(ustoz_id=ustoz_id)
    elif _is_ustoz(request.user) and teacher:
        qs = qs.filter(ustoz=teacher)
        ustoz_id = str(teacher.pk)
    elif _is_admin(request.user):
        if ustoz_id:
            qs = qs.filter(ustoz_id=ustoz_id)

    if hafta_turi in ('toq', 'juft'):
        qs = qs.filter(Q(hafta_turi='har_doim') | Q(hafta_turi=hafta_turi))

    darslar_by_kun = {kun: [] for kun, _ in Lesson.HAFTA_KUNI_CHOICES}
    for lesson in qs.order_by('boshlanish_vaqti'):
        darslar_by_kun[lesson.hafta_kuni].append(lesson)

    fanlar = list(Lesson.objects.values_list('fan_nomi', flat=True).distinct())
    rang_palette = [
        '#4e79a7', '#f28e2b', '#e15759', '#76b7b2', '#59a14f',
        '#edc948', '#b07aa1', '#ff9da7', '#9c755f', '#bab0ac',
    ]
    fan_ranglari = {fan: rang_palette[i % len(rang_palette)] for i, fan in enumerate(fanlar)}

    for lessons in darslar_by_kun.values():
        for lesson in lessons:
            lesson.rang = fan_ranglari.get(lesson.fan_nomi, '#6c757d')

    jadval = [
        {'key': kun, 'label': label, 'darslar': darslar_by_kun[kun]}
        for kun, label in Lesson.HAFTA_KUNI_CHOICES
    ]

    context = {
        'jadval': jadval,
        'ustozlar': Teacher.objects.all() if (_is_admin(request.user) or _is_bola(request.user)) else Teacher.objects.none(),
        'ustoz_id': ustoz_id,
        'hafta_turi': hafta_turi,
    }
    return render(request, 'schedule/schedule.html', context)


@login_required
def statistics_view(request):
    if _is_bola(request.user):
        return redirect('schedule')

    teacher = _get_teacher(request.user)

    if _is_admin(request.user):
        base_qs = Lesson.objects
        ustozlar_statistika = Teacher.objects.annotate(darslar_soni=Count('darslar')).order_by('-darslar_soni')
        top5_ustozlar = ustozlar_statistika[:5]
    else:
        base_qs = Lesson.objects.filter(ustoz=teacher) if teacher else Lesson.objects.none()
        ustozlar_statistika = None
        top5_ustozlar = None

    kunlar_statistika = []
    for kun_key, kun_label in Lesson.HAFTA_KUNI_CHOICES:
        soni = base_qs.filter(hafta_kuni=kun_key).count()
        kunlar_statistika.append({'kun': kun_label, 'soni': soni})

    fanlar_statistika = (
        base_qs.values('fan_nomi').annotate(soni=Count('id')).order_by('-soni')
    )

    context = {
        'ustozlar_statistika': ustozlar_statistika,
        'kunlar_statistika': kunlar_statistika,
        'fanlar_statistika': fanlar_statistika,
        'top5_ustozlar': top5_ustozlar,
    }
    return render(request, 'schedule/statistics.html', context)


@login_required
def dastur_view(request):
    if _is_bola(request.user):
        return redirect('schedule')
    return render(request, 'schedule/dastur.html')
